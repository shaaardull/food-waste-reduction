"""Tests for the restaurant-scoped bills list + monthly xlsx export.

Covers:
  • Cursor pagination on the list endpoint (50/page, next_cursor round-trip)
  • Status filter (paid / unpaid / voided / all)
  • Cross-restaurant staff → NOT_RESTAURANT_STAFF (403)
  • Export headers (Content-Type + Content-Disposition + filename slug)
  • Export against an empty month → still a valid, openable xlsx
  • Export cell values, currency format, TOTAL SUM formulas

The xlsx assertions load the response body back through openpyxl —
we don't just check bytes come back, we open the workbook and assert
on cell content, number formats, and the formula strings so a future
refactor that silently drops the sums fails a test.
"""
from __future__ import annotations

import io
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from urllib.parse import quote

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.bill import Bill
from app.models.meal_session import MealSession, MealSessionItem
from app.models.restaurant import Restaurant
from tests.conftest import (
    login,
    make_email,
    make_restaurant,
    make_staff,
    make_table_code,
)


# ── helpers ───────────────────────────────────────────────────────────


def _diner_user(db: Session) -> tuple[str, str]:
    from app.models.user import User
    from app.security import hash_password

    email = make_email("bills-export-diner")
    u = User(
        email=email,
        display_name="Bills Diner",
        role="diner",
        password_hash=hash_password("plate-clean-demo"),
    )
    db.add(u)
    db.flush()
    db.commit()
    return str(u.id), email


def _seed_bill(
    db: Session,
    *,
    restaurant_id,
    menu_items,
    when: datetime,
    diner_user_id: str,
    bill_number: str,
    paid: bool = False,
    voided: bool = False,
    voided_reason: str | None = None,
    entry_channel: str = "walkin",
    is_takeaway: bool = False,
    customer_email: str | None = None,
    customer_phone: str | None = None,
) -> tuple[MealSession, Bill]:
    """Manufacture a bill+session pair at a specific point in time.

    Bypasses the generation router entirely so we can seed backdated
    months and specific statuses without stepping through the whole
    diner flow. `when` is used for BOTH session.started_at and
    bill.issued_at + bill.created_at.
    """
    session_status = "paid" if paid else ("voided" if voided else "billed")
    session = MealSession(
        diner_user_id=diner_user_id,
        restaurant_id=restaurant_id,
        table_code=make_table_code("bex"),
        status=session_status,
        entry_channel=entry_channel,
        is_takeaway=is_takeaway,
        started_at=when,
        expires_at=when + timedelta(hours=4),
        customer_email=customer_email,
        customer_phone=customer_phone,
        paid_at=when if paid else None,
        voided_at=when if voided else None,
        voided_reason=voided_reason,
    )
    db.add(session)
    db.flush()
    subtotal_minor = 0
    line_items: list[dict] = []
    for m in menu_items:
        db.add(
            MealSessionItem(
                meal_session_id=session.id,
                menu_item_id=m.id,
                quantity=1,
                portion_size="regular",
            )
        )
        subtotal_minor += m.price_minor
        line_items.append(
            {
                "menu_item_id": str(m.id),
                "name": m.name,
                "quantity": 1,
                "portion_size": "regular",
                "price_minor": m.price_minor,
                "line_total_minor": m.price_minor,
            }
        )
    cgst = subtotal_minor * 25 // 1000
    sgst = cgst
    bill = Bill(
        meal_session_id=session.id,
        restaurant_id=restaurant_id,
        bill_number=bill_number,
        subtotal_minor=subtotal_minor,
        discount_minor=0,
        reward_redemption_code=None,
        taxable_amount_minor=subtotal_minor,
        cgst_rate=Decimal("0.025"),
        sgst_rate=Decimal("0.025"),
        cgst_amount_minor=cgst,
        sgst_amount_minor=sgst,
        total_minor=subtotal_minor + cgst + sgst,
        currency="INR",
        line_items_json=line_items,
        delivery_status="pending",
        issued_at=when,
        created_at=when,
        updated_at=when,
    )
    db.add(bill)
    db.commit()
    return session, bill


# ── list endpoint: pagination ─────────────────────────────────────────


@pytest.mark.asyncio
async def test_list_bills_paginates_at_fifty(client, db):
    restaurant, items, _ = make_restaurant(db, name="ListPage")
    staff = make_staff(db, restaurant.id)
    diner_id, _ = _diner_user(db)

    # Seed 60 bills — 55 in July 2026, 5 in June 2026 — so we can
    # verify month filtering strips out the June ones and pagination
    # walks the 55 July bills across two pages.
    july_base = datetime(2026, 7, 15, 12, 0, 0, tzinfo=UTC)
    june_base = datetime(2026, 6, 15, 12, 0, 0, tzinfo=UTC)
    for i in range(55):
        _seed_bill(
            db,
            restaurant_id=restaurant.id,
            menu_items=items[:1],
            when=july_base - timedelta(minutes=i),
            diner_user_id=diner_id,
            bill_number=f"JUL/{i:05d}",
        )
    for i in range(5):
        _seed_bill(
            db,
            restaurant_id=restaurant.id,
            menu_items=items[:1],
            when=june_base - timedelta(minutes=i),
            diner_user_id=diner_id,
            bill_number=f"JUN/{i:05d}",
        )

    token = await login(client, staff.email)

    # Page 1: 50 rows, cursor set.
    res = await client.get(
        f"/api/v1/restaurants/{restaurant.id}/bills?month=2026-07",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert len(body["rows"]) == 50
    assert body["next_cursor"] is not None
    # Sort: created_at DESC — first row is the most recent (i=0).
    assert body["rows"][0]["bill_number"] == "JUL/00000"

    # Page 2 with the returned cursor: remaining 5, no more.
    res2 = await client.get(
        f"/api/v1/restaurants/{restaurant.id}/bills"
        f"?month=2026-07&cursor={quote(body['next_cursor'])}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res2.status_code == 200, res2.text
    body2 = res2.json()
    assert len(body2["rows"]) == 5
    assert body2["next_cursor"] is None


# ── list endpoint: status filter ──────────────────────────────────────


@pytest.mark.asyncio
async def test_list_bills_status_filter(client, db):
    restaurant, items, _ = make_restaurant(db, name="StatusFlt")
    staff = make_staff(db, restaurant.id)
    diner_id, _ = _diner_user(db)

    base = datetime(2026, 7, 10, 10, 0, 0, tzinfo=UTC)
    # 2 paid, 1 unpaid, 1 voided
    _seed_bill(
        db,
        restaurant_id=restaurant.id,
        menu_items=items[:1],
        when=base,
        diner_user_id=diner_id,
        bill_number="STF/PAID-1",
        paid=True,
    )
    _seed_bill(
        db,
        restaurant_id=restaurant.id,
        menu_items=items[:1],
        when=base + timedelta(minutes=1),
        diner_user_id=diner_id,
        bill_number="STF/PAID-2",
        paid=True,
    )
    _seed_bill(
        db,
        restaurant_id=restaurant.id,
        menu_items=items[:1],
        when=base + timedelta(minutes=2),
        diner_user_id=diner_id,
        bill_number="STF/UNPAID",
    )
    _seed_bill(
        db,
        restaurant_id=restaurant.id,
        menu_items=items[:1],
        when=base + timedelta(minutes=3),
        diner_user_id=diner_id,
        bill_number="STF/VOIDED",
        voided=True,
        voided_reason="wrong table",
    )

    token = await login(client, staff.email)

    async def _fetch(status: str) -> list[dict]:
        r = await client.get(
            f"/api/v1/restaurants/{restaurant.id}/bills"
            f"?month=2026-07&status={status}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200, r.text
        return r.json()["rows"]

    assert len(await _fetch("all")) == 4
    paid_rows = await _fetch("paid")
    assert len(paid_rows) == 2
    assert all(r["status"] == "paid" for r in paid_rows)

    unpaid_rows = await _fetch("unpaid")
    assert len(unpaid_rows) == 1
    assert unpaid_rows[0]["bill_number"] == "STF/UNPAID"

    voided_rows = await _fetch("voided")
    assert len(voided_rows) == 1
    assert voided_rows[0]["voided_reason"] == "wrong table"


# ── list endpoint: cross-restaurant staff guard ───────────────────────


@pytest.mark.asyncio
async def test_list_bills_cross_restaurant_staff_forbidden(client, db):
    """Staff of restaurant A hitting restaurant B → NOT_RESTAURANT_STAFF."""
    r_a, _items_a, _ = make_restaurant(db, name="RA")
    r_b, _items_b, _ = make_restaurant(db, name="RB")
    staff_a = make_staff(db, r_a.id)
    token = await login(client, staff_a.email)

    res = await client.get(
        f"/api/v1/restaurants/{r_b.id}/bills?month=2026-07",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 403
    body = res.json()
    assert body["error"]["code"] == "NOT_RESTAURANT_STAFF"


# ── export: headers + filename ────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_returns_xlsx_with_attachment_header(client, db):
    restaurant, items, _ = make_restaurant(db, name="ExportHdr")
    staff = make_staff(db, restaurant.id)
    diner_id, _ = _diner_user(db)

    _seed_bill(
        db,
        restaurant_id=restaurant.id,
        menu_items=items[:1],
        when=datetime(2026, 7, 4, 13, 30, tzinfo=UTC),
        diner_user_id=diner_id,
        bill_number="EXP/HDR-1",
        paid=True,
    )
    token = await login(client, staff.email)
    res = await client.get(
        f"/api/v1/restaurants/{restaurant.id}/bills/export?month=2026-07",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200, res.text
    assert (
        res.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disp = res.headers["content-disposition"]
    assert "attachment" in disp
    assert f"plateclean-{restaurant.slug}-bills-2026-07.xlsx" in disp


# ── export: empty month still valid ───────────────────────────────────


@pytest.mark.asyncio
async def test_export_empty_month_returns_valid_xlsx(client, db):
    """A month with zero bills must still stream a well-formed xlsx —
    the CA opens it and sees the restaurant header + a note explaining
    the period was empty, not a broken download."""
    restaurant, _items, _ = make_restaurant(db, name="ExportEmpty")
    staff = make_staff(db, restaurant.id)
    token = await login(client, staff.email)

    res = await client.get(
        f"/api/v1/restaurants/{restaurant.id}/bills/export?month=2026-01",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws is not None
    # Header rows present.
    assert ws.cell(row=1, column=1).value == restaurant.name
    # Row 6 (first data row) carries the "no bills" note when the
    # month is empty.
    note = ws.cell(row=6, column=1).value
    assert isinstance(note, str) and "No bills" in note


# ── export: cell values + TOTAL formulas ──────────────────────────────


@pytest.mark.asyncio
async def test_export_cell_values_and_total_formulas(client, db):
    restaurant, items, _ = make_restaurant(db, name="ExportCells")
    # GSTIN + explicit address so the row-2 metadata renders both.
    r_db = db.get(Restaurant, restaurant.id)
    assert r_db is not None
    r_db.gstin = "27AABCU9603R1ZM"
    r_db.address = "12 Chapel St, Bandra West, Mumbai"
    db.commit()

    staff = make_staff(db, restaurant.id)
    diner_id, _ = _diner_user(db)

    base = datetime(2026, 7, 4, 13, 30, tzinfo=UTC)
    seeded: list[tuple[MealSession, Bill]] = []
    for i in range(5):
        seeded.append(
            _seed_bill(
                db,
                restaurant_id=restaurant.id,
                menu_items=items[:1],
                when=base + timedelta(hours=i),
                diner_user_id=diner_id,
                bill_number=f"EXC/{i:05d}",
                paid=(i % 2 == 0),
            )
        )

    token = await login(client, staff.email)
    res = await client.get(
        f"/api/v1/restaurants/{restaurant.id}/bills/export?month=2026-07",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws is not None

    # Sheet title reflects the readable month.
    assert ws.title == "Bills — July 2026"

    # Row 1: restaurant name (merged). Row 2: metadata line contains
    # both GSTIN and address. Row 3: report period spelt out.
    assert ws.cell(row=1, column=1).value == restaurant.name
    row2 = ws.cell(row=2, column=1).value
    assert "27AABCU9603R1ZM" in row2
    assert "Bandra West" in row2
    row3 = ws.cell(row=3, column=1).value
    assert "01 Jul 2026" in row3
    assert "31 Jul 2026" in row3

    # Column headers on row 5.
    expected_headers = [
        "Date",
        "Bill Number",
        "Channel",
        "Table",
        "Customer Email",
        "Customer Phone",
        "Items",
        "Subtotal (₹)",
        "CGST %",
        "CGST (₹)",
        "SGST %",
        "SGST (₹)",
        "Total (₹)",
        "Status",
    ]
    for col_idx, label in enumerate(expected_headers, start=1):
        assert ws.cell(row=5, column=col_idx).value == label

    # Data rows: 5 bills, ascending by created_at (chronological).
    # First data row is 6, last is 10.
    first_data_row = 6
    last_data_row = 10
    for offset in range(5):
        row = first_data_row + offset
        assert ws.cell(row=row, column=2).value == f"EXC/{offset:05d}"
        # Money cells are floats with the ₹#,##0.00 format.
        subtotal = ws.cell(row=row, column=8)
        assert subtotal.number_format == "₹#,##0.00"
        assert isinstance(subtotal.value, (int, float))
        # cgst_amt + sgst_amt + subtotal ≈ total (all in rupees).
        cgst_amt = ws.cell(row=row, column=10).value
        sgst_amt = ws.cell(row=row, column=12).value
        total = ws.cell(row=row, column=13).value
        assert isinstance(cgst_amt, (int, float))
        assert isinstance(sgst_amt, (int, float))
        assert abs((subtotal.value + cgst_amt + sgst_amt) - total) < 0.01

    # TOTAL row: one blank row after last data row, then bold TOTAL +
    # SUM formulas over the money columns.
    total_row = last_data_row + 2
    assert ws.cell(row=total_row, column=1).value == "TOTAL"
    for money_col, letter in ((8, "H"), (10, "J"), (12, "L"), (13, "M")):
        formula = ws.cell(row=total_row, column=money_col).value
        assert isinstance(formula, str)
        assert formula.startswith(f"=SUM({letter}{first_data_row}:{letter}{last_data_row}")

    # Panes frozen from row 6 so the header rows stay visible on scroll.
    assert ws.freeze_panes == "A6"

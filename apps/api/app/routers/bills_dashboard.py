"""Restaurant-scoped bills list + CA-facing monthly Excel export.

Two READ paths only:
  GET /restaurants/{id}/bills           — paginated list, 50/page
  GET /restaurants/{id}/bills/export    — xlsx download, one month

Everything about creating / voiding / emailing bills stays on the
session and legacy /bills routers. This module never mutates.

Status derivation. Bills don't carry paid/voided state — the parent
`meal_sessions` row does (walk-ins get `paid_at` / `voided_at` set by
POS-side actions, and QR sessions never enter either state under the
current phase 1 flows). So we derive:
  • voided  ← meal_sessions.voided_at IS NOT NULL
  • paid    ← voided_at IS NULL AND meal_sessions.paid_at IS NOT NULL
  • unpaid  ← neither
That mirrors the `mark-paid` semantic from apps/api/app/routers/sessions.py.
"""
from __future__ import annotations

import io
from calendar import monthrange
from datetime import UTC, datetime
from decimal import Decimal
from typing import Any, Literal
from uuid import UUID
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db
from app.errors import NotRestaurantStaff
from app.models.bill import Bill
from app.models.meal_session import MealSession, MealSessionItem
from app.models.menu_item import MenuItem
from app.models.restaurant import Restaurant, RestaurantStaff
from app.models.user import User
from app.security import get_current_user

router = APIRouter()


StatusFilter = Literal["paid", "unpaid", "voided", "all"]


# ── auth ──────────────────────────────────────────────────────────────


async def _ensure_staff(
    db: AsyncSession, user: User, restaurant_id: UUID
) -> None:
    """Flat any-staff check (owner/manager/server) mirroring dashboard.py.

    Cross-restaurant staff → NOT_RESTAURANT_STAFF envelope the frontend
    already knows how to route to the "not on staff" screen.
    """
    if user.role == "admin":
        return
    if user.role != "staff":
        raise NotRestaurantStaff()
    res = await db.execute(
        select(RestaurantStaff).where(
            RestaurantStaff.user_id == user.id,
            RestaurantStaff.restaurant_id == restaurant_id,
        )
    )
    if res.scalar_one_or_none() is None:
        raise NotRestaurantStaff()


# ── time helpers ──────────────────────────────────────────────────────


def _tz_for(restaurant: Restaurant) -> ZoneInfo:
    """Resolve the restaurant's timezone, falling back to UTC on a
    corrupt / stale IANA string rather than 500-ing the endpoint."""
    try:
        return ZoneInfo(restaurant.timezone or "UTC")
    except ZoneInfoNotFoundError:
        return ZoneInfo("UTC")


def _parse_month(month_str: str) -> tuple[int, int]:
    """Parse `YYYY-MM` into (year, month). Raises 400 on anything else."""
    try:
        year_s, mon_s = month_str.split("-")
        year = int(year_s)
        mon = int(mon_s)
        if not (1 <= mon <= 12):
            raise ValueError
        # Reasonable bounds — pre-2000 or post-2100 is almost certainly
        # a typo, and letting them through would make the sheet header
        # ("Report period: 01 Jan 2087 – …") read as garbage.
        if not (2000 <= year <= 2100):
            raise ValueError
    except (ValueError, AttributeError) as exc:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_MONTH",
                "message": "month must be YYYY-MM.",
            },
        ) from exc
    return year, mon


def _month_bounds_utc(
    year: int, month: int, tz: ZoneInfo
) -> tuple[datetime, datetime]:
    """Return (start, end) as UTC datetimes covering [month, next-month)
    in the restaurant's local timezone. Half-open so a bill created at
    exactly 00:00:00 on the 1st of the next month falls into the next
    period, not this one."""
    _, last_day = monthrange(year, month)  # noqa: F841 — kept for clarity
    start_local = datetime(year, month, 1, 0, 0, 0, tzinfo=tz)
    if month == 12:
        end_local = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=tz)
    else:
        end_local = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=tz)
    return start_local.astimezone(UTC), end_local.astimezone(UTC)


def _current_month_in_tz(tz: ZoneInfo) -> tuple[int, int]:
    now_local = datetime.now(tz)
    return now_local.year, now_local.month


# ── shared row shape ──────────────────────────────────────────────────


def _derive_status(session: MealSession) -> str:
    """See module docstring for the rule. Voided beats paid beats unpaid.
    Kept as a pure function so the export path can reuse it without
    re-implementing the ladder."""
    if session.voided_at is not None:
        return "voided"
    if session.paid_at is not None:
        return "paid"
    return "unpaid"


def _bill_row(bill: Bill, session: MealSession) -> dict[str, Any]:
    # gst_rate on the response is the total (CGST + SGST) so the CA-side
    # summary matches the ledger — the CGST/SGST split lives on the
    # export sheet columns, not on the JSON row where callers just want
    # the effective rate.
    gst_rate = (bill.cgst_rate or Decimal(0)) + (bill.sgst_rate or Decimal(0))
    gst_amount = bill.cgst_amount_minor + bill.sgst_amount_minor
    return {
        "id": str(bill.id),
        "meal_session_id": str(bill.meal_session_id),
        "bill_number": bill.bill_number,
        "created_at": bill.created_at.astimezone(UTC).isoformat(),
        "channel": session.entry_channel,
        "is_takeaway": session.is_takeaway,
        "table_code": session.table_code,
        "customer_email": session.customer_email,
        "customer_phone": session.customer_phone,
        "subtotal_minor": bill.subtotal_minor,
        "gst_rate": str(gst_rate),
        "gst_amount_minor": gst_amount,
        "total_minor": bill.total_minor,
        "status": _derive_status(session),
        "voided_at": (
            session.voided_at.astimezone(UTC).isoformat()
            if session.voided_at
            else None
        ),
        "voided_reason": session.voided_reason,
    }


# ── list endpoint ─────────────────────────────────────────────────────


@router.get("/{restaurant_id}/bills")
async def list_bills(
    restaurant_id: UUID,
    month: str | None = Query(default=None),
    status: StatusFilter = Query(default="all"),
    limit: int = Query(default=50, ge=1, le=50),
    cursor: str | None = Query(default=None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    """Paginated bills list for the staff dashboard's Bills view.

    Sort: `Bill.created_at DESC`. Cursor is the last-seen row's
    `created_at` ISO8601 — the next page is `created_at < cursor`.
    """
    await _ensure_staff(db, user, restaurant_id)
    restaurant = await db.get(Restaurant, restaurant_id)
    if restaurant is None:
        raise HTTPException(status_code=404, detail="Restaurant not found")

    tz = _tz_for(restaurant)
    if month:
        year, mon = _parse_month(month)
    else:
        year, mon = _current_month_in_tz(tz)
    start_utc, end_utc = _month_bounds_utc(year, mon, tz)

    # Cursor: keyset over Bill.created_at. Bills within the same second
    # sort deterministically by id (implicit tie-breaker for the rare
    # burst-insert case).
    cursor_dt: datetime | None = None
    if cursor:
        try:
            cursor_dt = datetime.fromisoformat(cursor)
            if cursor_dt.tzinfo is None:
                cursor_dt = cursor_dt.replace(tzinfo=UTC)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail={
                    "code": "INVALID_CURSOR",
                    "message": "cursor must be ISO8601.",
                },
            ) from exc

    stmt = (
        select(Bill, MealSession)
        .join(MealSession, MealSession.id == Bill.meal_session_id)
        .where(
            Bill.restaurant_id == restaurant_id,
            Bill.created_at >= start_utc,
            Bill.created_at < end_utc,
        )
    )
    if status == "voided":
        stmt = stmt.where(MealSession.voided_at.is_not(None))
    elif status == "paid":
        stmt = stmt.where(
            MealSession.voided_at.is_(None),
            MealSession.paid_at.is_not(None),
        )
    elif status == "unpaid":
        stmt = stmt.where(
            MealSession.voided_at.is_(None),
            MealSession.paid_at.is_(None),
        )
    if cursor_dt is not None:
        stmt = stmt.where(Bill.created_at < cursor_dt)

    # Fetch limit+1 so we can tell whether a next page exists without a
    # second COUNT round-trip.
    stmt = stmt.order_by(Bill.created_at.desc(), Bill.id.desc()).limit(limit + 1)

    result = await db.execute(stmt)
    pairs = result.all()

    has_more = len(pairs) > limit
    if has_more:
        pairs = pairs[:limit]

    rows = [_bill_row(bill, session) for (bill, session) in pairs]
    next_cursor: str | None = None
    if has_more and pairs:
        last_bill, _ = pairs[-1]
        next_cursor = last_bill.created_at.astimezone(UTC).isoformat()

    return {"rows": rows, "next_cursor": next_cursor}


# ── excel export ──────────────────────────────────────────────────────


_MONTH_NAMES = (
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
)


def _fmt_local(dt: datetime, tz: ZoneInfo) -> datetime:
    """openpyxl writes tz-naive datetimes; convert to the restaurant's
    local wall-clock time so the CA sees dates in the timezone the
    restaurant lives in, not UTC."""
    return dt.astimezone(tz).replace(tzinfo=None)


def _autofit_column_widths(ws, ncols: int, cap: int = 40) -> None:
    """Cheap autofit — walk each column, take max(len(str(cell.value))),
    cap at `cap` chars so a long email doesn't blow the layout. Not
    perfect (Excel measures pixels, not chars) but close enough for a
    CA-facing sheet."""
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                val = cell.value
                # Datetimes render wider than str(datetime) — pad a bit.
                if isinstance(val, datetime):
                    length = 18
                else:
                    length = len(str(val))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[letter].width = min(cap, max(10, max_len + 2))


def _channel_label(channel: str, is_takeaway: bool) -> str:
    if channel == "qr":
        return "QR"
    if is_takeaway:
        return "Takeaway"
    return "Walk-in"


def _items_summary(items_rows: list[tuple[MealSessionItem, MenuItem]]) -> str:
    """Semi-colon separated `N × name` list. Empty string if no items."""
    parts: list[str] = []
    for msi, menu in items_rows:
        parts.append(f"{msi.quantity} × {menu.name}")
    return "; ".join(parts)


async def _fetch_month_bills(
    db: AsyncSession,
    *,
    restaurant_id: UUID,
    start_utc: datetime,
    end_utc: datetime,
) -> list[tuple[Bill, MealSession]]:
    """Ascending by issued_at — chronological order is what the CA
    expects on a monthly ledger."""
    stmt = (
        select(Bill, MealSession)
        .join(MealSession, MealSession.id == Bill.meal_session_id)
        .where(
            Bill.restaurant_id == restaurant_id,
            Bill.created_at >= start_utc,
            Bill.created_at < end_utc,
        )
        .order_by(Bill.created_at.asc())
    )
    return list((await db.execute(stmt)).all())


async def _fetch_items_by_session(
    db: AsyncSession, session_ids: list[UUID]
) -> dict[UUID, list[tuple[MealSessionItem, MenuItem]]]:
    if not session_ids:
        return {}
    stmt = (
        select(MealSessionItem, MenuItem)
        .join(MenuItem, MealSessionItem.menu_item_id == MenuItem.id)
        .where(MealSessionItem.meal_session_id.in_(session_ids))
    )
    out: dict[UUID, list[tuple[MealSessionItem, MenuItem]]] = {}
    for msi, menu in (await db.execute(stmt)).all():
        out.setdefault(msi.meal_session_id, []).append((msi, menu))
    return out


def _build_workbook(
    *,
    restaurant: Restaurant,
    year: int,
    month: int,
    tz: ZoneInfo,
    pairs: list[tuple[Bill, MealSession]],
    items_by_session: dict[UUID, list[tuple[MealSessionItem, MenuItem]]],
) -> bytes:
    """Assemble the xlsx workbook in memory and return the bytes.

    Layout (CA-ready):
      Row 1  restaurant name (bold ~16pt, merged A1:N1)
      Row 2  GSTIN · address (~11pt muted, merged A2:N2)
      Row 3  report period (merged A3:N3)
      Row 4  blank
      Row 5  bold column headers (frozen)
      Row 6+ data
      TOTAL  bold, one blank row before, live SUM formulas
    """
    month_name = _MONTH_NAMES[month - 1]
    wb = Workbook()
    ws = wb.active
    if ws is None:
        # openpyxl always creates one sheet, but the type is Optional.
        ws = wb.create_sheet()
    ws.title = f"Bills — {month_name} {year}"

    ncols = 14  # A..N

    # ── header rows ──
    name_font = Font(bold=True, size=16)
    meta_font = Font(size=11, color="666666")
    period_font = Font(size=11, italic=True, color="444444")

    ws.cell(row=1, column=1, value=restaurant.name).font = name_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    meta_parts: list[str] = []
    if restaurant.gstin:
        meta_parts.append(f"GSTIN: {restaurant.gstin}")
    meta_parts.append(restaurant.address)
    ws.cell(row=2, column=1, value=" · ".join(meta_parts)).font = meta_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    _, last_day = monthrange(year, month)
    period = (
        f"Report period: 01 {month_name[:3]} {year} "
        f"– {last_day:02d} {month_name[:3]} {year}"
    )
    ws.cell(row=3, column=1, value=period).font = period_font
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)

    # Row 4 is a spacer — leave blank.

    # ── column headers ──
    headers = [
        "Date",
        "Bill Number",
        "Channel",
        "Table",
        "Customer Email",
        "Customer Phone",
        "Items",
        "Subtotal (₹)",
        "CGST %",
        "CGST (₹)",
        "SGST %",
        "SGST (₹)",
        "Total (₹)",
        "Status",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F2A24", end_color="1F2A24", fill_type="solid"
    )
    header_row = 5
    for i, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left" if i not in (8, 10, 12, 13) else "right")

    # Freeze rows 1..5 so scrolling data keeps the header + metadata
    # visible.
    ws.freeze_panes = f"A{header_row + 1}"

    currency_fmt = "₹#,##0.00"
    date_fmt = "dd mmm yyyy, hh:mm"

    # ── data rows ──
    first_data_row = header_row + 1
    if not pairs:
        # Empty-month sheet — one placeholder row so the file still
        # opens with content and isn't confused for an error.
        ws.cell(row=first_data_row, column=1, value="No bills recorded for this period.").font = Font(
            italic=True, color="888888"
        )
        ws.merge_cells(
            start_row=first_data_row,
            start_column=1,
            end_row=first_data_row,
            end_column=ncols,
        )
        _autofit_column_widths(ws, ncols)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    row_idx = first_data_row
    for bill, session in pairs:
        items = items_by_session.get(session.id, [])
        subtotal = bill.subtotal_minor / 100
        cgst_rate = float(bill.cgst_rate or 0) * 100
        sgst_rate = float(bill.sgst_rate or 0) * 100
        cgst_amt = bill.cgst_amount_minor / 100
        sgst_amt = bill.sgst_amount_minor / 100
        total = bill.total_minor / 100

        status = _derive_status(session)
        if status == "voided":
            status_label = "Voided"
            if session.voided_reason:
                status_label = f"Voided ({session.voided_reason})"
        elif status == "paid":
            status_label = "Paid"
        else:
            status_label = "Unpaid"

        table_display = "—" if session.is_takeaway else session.table_code

        row_values: list[Any] = [
            _fmt_local(bill.created_at, tz),
            bill.bill_number,
            _channel_label(session.entry_channel, session.is_takeaway),
            table_display,
            session.customer_email or "",
            session.customer_phone or "",
            _items_summary(items),
            subtotal,
            cgst_rate,
            cgst_amt,
            sgst_rate,
            sgst_amt,
            total,
            status_label,
        ]
        for col_idx, val in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

        # Formats — apply per-cell so a single money column with a
        # blank value doesn't inherit an unrelated format.
        ws.cell(row=row_idx, column=1).number_format = date_fmt
        for money_col in (8, 10, 12, 13):
            ws.cell(row=row_idx, column=money_col).number_format = currency_fmt
        for pct_col in (9, 11):
            ws.cell(row=row_idx, column=pct_col).number_format = "0.00"

        row_idx += 1

    last_data_row = row_idx - 1

    # ── TOTAL row ──
    # Skip one blank row after data, then bold TOTAL + SUM formulas.
    total_row = last_data_row + 2
    total_font = Font(bold=True)
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    for money_col in (8, 10, 12, 13):
        letter = get_column_letter(money_col)
        formula = f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
        c = ws.cell(row=total_row, column=money_col, value=formula)
        c.font = total_font
        c.number_format = currency_fmt

    _autofit_column_widths(ws, ncols)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@router.get("/{restaurant_id}/bills/export")
async def export_bills_xlsx(
    restaurant_id: UUID,
    month: str = Query(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Stream a CA-ready xlsx of every bill for the given month."""
    await _ensure_staff(db, user, restaurant_id)
    restaurant = await db.get(Restaurant, restaurant_id)
    if restaurant is None:
        raise HTTPException(status_code=404, detail="Restaurant not found")

    year, mon = _parse_month(month)
    tz = _tz_for(restaurant)
    start_utc, end_utc = _month_bounds_utc(year, mon, tz)

    pairs = await _fetch_month_bills(
        db,
        restaurant_id=restaurant_id,
        start_utc=start_utc,
        end_utc=end_utc,
    )
    session_ids = [session.id for (_, session) in pairs]
    items_by_session = await _fetch_items_by_session(db, session_ids)

    xlsx_bytes = _build_workbook(
        restaurant=restaurant,
        year=year,
        month=mon,
        tz=tz,
        pairs=pairs,
        items_by_session=items_by_session,
    )

    filename = f"plateclean-{restaurant.slug}-bills-{year:04d}-{mon:02d}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type=_XLSX_MIME,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
